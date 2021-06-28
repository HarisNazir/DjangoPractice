from django.http.response import HttpResponse
from django.http.response import FileResponse
from openpyxl import Workbook, load_workbook
from django.http import FileResponse
from .models import MonthlyDataUpload

def AddToModel(response):
    wb = load_workbook(response)
    ws = wb.active
    male_beneficiaries = ws['C3']
    female_beneficiaries = ws['C4']
    cash_value = ws['C5']
    country_iso = ws['A1']
    
    return MonthlyDataUpload.objects.create(
        beneficiaries_assisted_males = 'male_beneficiaries',
        beneficiaries_assisted_females = female_beneficiaries,
        cash_value = 'cash_value',
        country_iso = 'country_iso3'

    )
