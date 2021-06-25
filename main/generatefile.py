from django.http.response import FileResponse
from openpyxl import Workbook
from django.http import FileResponse


def generateExcelFile(data):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = data.get('alpha3code')
    country_name = data.get('name')
    ws['B1'] = (f'Monthly data for {country_name}')
    ws['B3'] = ('Beneficiaries assisted (Male):')
    ws['B4'] = ('Beneficiaries assisted (Female):')
    currency = data.get('currency'[0])
    ws['B5'] = (f'Cash value of assistance {currency}:')

    wb.save("country.xlsx")

def download_file():
    response = FileResponse(open('country.xlsx', 'rb'))
    return response